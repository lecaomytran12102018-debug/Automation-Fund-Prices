"""
Cài đặt trước khi chạy:
    pip install playwright pandas openpyxl
    playwright install chromium

"""

import json
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError

# ----------------------------------------------------------------------------
# CẤU HÌNH
# ----------------------------------------------------------------------------

# Mã quỹ (fundId) -> Tên cột hiển thị trong Excel.
# Sửa lại tên hiển thị bên phải cho đúng tên quỹ thật nếu cần.
FUND_NAME_MAP = {
    "VNAGR": "Quỹ Tăng Trưởng",
    "VNGRW": "Quỹ Phát Triển",
    "VNBAL": "Quỹ Cân Bằng",
    "VNDIV": "Quỹ Ổn Định",
    "VNFIX": "Quỹ Tích Lũy",
    "VNMMK": "Quỹ Bảo Toàn",
    "VN035": "Quỹ Hưng Thịnh 2035",
    "VN040": "Quỹ Hưng Thịnh 2040",
    "VN045": "Quỹ Hưng Thịnh 2045",
    "VNTCM": "Quỹ Manulink Tiền Linh Hoạt",
    "VNTCF": "Quỹ Manulink Trái Phiếu",
    "VNTCE": "Quỹ Manulink Cổ Phiếu",
}

# Đường dẫn workbook tổng hợp cần cập nhật (sửa lại path đúng máy bạn).
INPUT_WORKBOOK = Path("gia_don_vi_quy_TONGHOP.xlsx")
OUTPUT_WORKBOOK = INPUT_WORKBOOK
SHEET_NAME = "Manulife"

BASE_URL = "https://portal.manulife.com.vn/cdn/mkt/don-vi-quy/index.html#/detail/{fund_id}"
API_PATH_FRAGMENT = "/s/bff/fund/prices"

HEADLESS = False  # để False cho dễ qua Akamai; đổi True nếu chạy trên server không có màn hình


# ----------------------------------------------------------------------------
# LẤY DỮ LIỆU TỪ MANULIFE (PLAYWRIGHT)
# ----------------------------------------------------------------------------

def fetch_fund_history(fund_id: str, playwright):
    """Mở trang quỹ, bấm nút khoảng thời gian dài nhất, bắt response API."""
    browser = playwright.chromium.launch(headless=HEADLESS, channel="chrome")
    context = browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
        ),
        locale="vi-VN",
    )
    page = context.new_page()

    captured = {"all": []}

    def handle_response(response):
        if API_PATH_FRAGMENT in response.url and response.request.method == "POST":
            try:
                body = response.json()
            except Exception:
                body = response.text()
            captured["all"].append({"status": response.status, "body": body})

    page.on("response", handle_response)

    url = BASE_URL.format(fund_id=fund_id)
    print(f"[{fund_id}] Đang mở: {url}")
    page.goto(url, wait_until="networkidle", timeout=60000)

    # Chờ Akamai sensor script chạy xong trước khi tương tác
    page.wait_for_timeout(6000)

    # Số response bắt được TRƯỚC khi bấm nút (thường là 1: load mặc định)
    responses_before_click = len(captured["all"])

    clicked = False
    for label in ("Từ ngày thành lập", "Tu ngay thanh lap"):
        try:
            page.get_by_text(label, exact=True).click(timeout=8000)
            clicked = True
            print(f"[{fund_id}] Đã bấm nút '{label}', đang chờ dữ liệu mới...")
            break
        except PWTimeoutError:
            continue

    if not clicked:
        print(f"[{fund_id}] KHÔNG bấm được nút 'Từ ngày thành lập' — "
              f"dữ liệu có thể chỉ là mặc định (1 năm).")

    # Chờ cho đến khi có response MỚI xuất hiện sau khi bấm nút
    deadline = time.time() + 20
    while len(captured["all"]) <= responses_before_click and time.time() < deadline:
        page.wait_for_timeout(500)

    # Đợi thêm chút để chắc chắn response cuối đã về đầy đủ
    page.wait_for_timeout(1500)

    context.close()
    browser.close()

    return captured["all"]


def get_sell_prices(fund_id: str, playwright) -> dict:
    """Trả về dict {datetime.date -> sellPrice} cho 1 quỹ.

    Có thể bắt được nhiều response (1 lúc load trang mặc định, 1 sau khi
    bấm nút "Từ ngày thành lập"). Chọn response có NHIỀU điểm dữ liệu
    nhất (tức là khoảng thời gian dài nhất) để đảm bảo lấy được full lịch
    sử, không bị dính bản mặc định (1 năm).
    """
    responses = fetch_fund_history(fund_id, playwright)

    if not responses:
        print(f"[{fund_id}] KHÔNG bắt được response nào từ API.")
        return {}

    ok_responses = [r for r in responses if r["status"] == 200 and isinstance(r["body"], dict)]
    if not ok_responses:
        print(f"[{fund_id}] Không có response 200 hợp lệ. Status thấy được: "
              f"{[r['status'] for r in responses]}")
        return {}

    # Chọn response có historyValues dài nhất = full lịch sử
    best = max(ok_responses, key=lambda r: len(r["body"].get("historyValues", [])))
    print(f"[{fund_id}] Bắt được {len(responses)} response, chọn response có "
          f"{len(best['body'].get('historyValues', []))} điểm dữ liệu (dài nhất).")

    history = best["body"].get("historyValues", [])

    prices = {}
    for item in history:
        try:
            d = datetime.strptime(item["date"], "%d/%m/%Y").date()
            prices[d] = item["sellPrice"]
        except (KeyError, ValueError):
            continue

    if prices:
        print(f"[{fund_id}] Lấy được {len(prices)} điểm, từ {min(prices)} đến {max(prices)}.")
    return prices


def fetch_all_funds() -> pd.DataFrame:
    """Gọi lần lượt 12 quỹ, trả về 1 DataFrame: cột Ngày + 1 cột mỗi quỹ."""
    all_series = {}

    with sync_playwright() as p:
        for fund_id in FUND_NAME_MAP:
            prices = get_sell_prices(fund_id, p)
            all_series[FUND_NAME_MAP[fund_id]] = prices

    # Hợp nhất tất cả các ngày xuất hiện ở bất kỳ quỹ nào
    all_dates = sorted(
        {d for series in all_series.values() for d in series.keys()},
        reverse=True,  # mới nhất lên đầu, giống các sheet khác
    )

    rows = []
    for d in all_dates:
        row = {"Ngày": d}
        for col_name, series in all_series.items():
            row[col_name] = series.get(d)
        rows.append(row)

    df = pd.DataFrame(rows, columns=["Ngày"] + list(FUND_NAME_MAP.values()))
    return df


# ----------------------------------------------------------------------------
# GHI VÀO WORKBOOK, GIỮ ĐÚNG ĐỊNH DẠNG CÁC SHEET KHÁC
# ----------------------------------------------------------------------------

def write_to_workbook(df: pd.DataFrame):
    wb = load_workbook(INPUT_WORKBOOK)

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME, 0)  # đặt lên đầu, giống vị trí hiện tại

    header_font = Font(name="Calibri", size=11, bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    body_font = Font(name="Calibri", size=11)
    number_format = "#,##0.000"  # giá Manulife có phần lẻ (vd 35,433.171)
    date_format = "dd/mm/yyyy"

    # Header
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = header_align

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            if col_idx == 1:
                cell.number_format = date_format
            else:
                cell.number_format = number_format

    # Độ rộng cột cho dễ đọc
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    wb.save(OUTPUT_WORKBOOK)
    print(f"\nĐã ghi xong. File kết quả: {OUTPUT_WORKBOOK.resolve()}")


# ----------------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------------

def main():
    df = fetch_all_funds()

    if df.empty:
        print("Không lấy được dữ liệu nào, dừng lại.")
        return

    print(f"\nTổng số ngày (hợp nhất tất cả các quỹ): {len(df)}")
    print(df.head())

    write_to_workbook(df)


if __name__ == "__main__":
    main()