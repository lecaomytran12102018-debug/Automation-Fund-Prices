# -*- coding: utf-8 -*-

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


DATE_FORMAT = "%d/%m/%Y"


def _parse_date(value):
    """Chuyển cell ngày (string hoặc datetime) về object datetime để so sánh/sort."""
    if isinstance(value, datetime):
        return value
    if value is None:
        return None
    value = str(value).strip()
    for fmt in (DATE_FORMAT, "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def update_fund_excel(
    file_path,
    new_data,
    sheet_name=None,
    date_col_name="Ngày",
    date_format=DATE_FORMAT,
    create_missing_columns=True,
    verbose=True,
):
    """
    Cập nhật các dòng giá mới vào file Excel raw data, không tạo duplicate ngày,
    không overlap dữ liệu cũ.

    Params
    ------
    file_path : str
        Đường dẫn file .xlsx cần update (file đã có sẵn header + data lịch sử).
    new_data : dict
        { "dd/mm/yyyy": { "Tên cột quỹ A": giá, "Tên cột quỹ B": giá, ... }, ... }
        Có thể đưa nhiều ngày 1 lần (vd chạy bù vài tuần).
    sheet_name : str hoặc None
        Tên sheet cần update. None = sheet active đầu tiên.
    date_col_name : str
        Tên cột chứa ngày trong header (mặc định "Ngày").
    date_format : str
        Format ghi ngày vào cell (mặc định dd/mm/yyyy).
    create_missing_columns : bool
        Nếu fund mới chưa có cột trong file -> tự tạo thêm cột cuối cùng.
    verbose : bool
        In log các dòng đã thêm / đã skip.

    Returns
    -------
    int : số dòng mới đã được thêm vào file.
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_font = Font(name="Calibri", size=11, bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    body_font = Font(name="Calibri", size=11)
    number_format = "#,##0"
    date_excel_format = "dd/mm/yyyy"

    # 1. Đọc header -> map tên cột -> số thứ tự cột
    header_row = 1
    headers = {}
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        val = cell.value

        if val is not None:
            headers[str(val).strip()] = col_idx

        # APPLY FORMAT
        cell.font = header_font
        header_align = Alignment(horizontal="center", vertical="center")


    if date_col_name not in headers:
        raise ValueError(
            f"Không tìm thấy cột '{date_col_name}' ở header của '{file_path}'."
        )
    date_col_idx = headers[date_col_name]

    # 2. Tạo thêm cột mới cho fund chưa từng có trong file (nếu được phép)
    all_fund_names = set()
    for row_data in new_data.values():
        all_fund_names.update(row_data.keys())

    missing_funds = [f for f in all_fund_names if f not in headers]
    if missing_funds:
        if not create_missing_columns:
            raise ValueError(
                f"Các cột chưa tồn tại trong file: {missing_funds}. "
                f"Set create_missing_columns=True nếu muốn tự tạo."
            )
        next_col = ws.max_column + 1
        for fund_name in missing_funds:
            cell = ws.cell(row=header_row, column=next_col, value=fund_name)
            cell.font = header_font
            cell.alignment = header_align
            headers[fund_name] = next_col
            next_col += 1
        if verbose:
            print(f"[+] Đã tạo cột mới: {missing_funds}")

    # 3. Lấy tập các ngày đã tồn tại trong file (để tránh duplicate)
    existing_dates = set()
    for row_idx in range(header_row + 1, ws.max_row + 1):
        d = _parse_date(ws.cell(row=row_idx, column=date_col_idx).value)
        if d:
            existing_dates.add(d.date())

    # 4. Lọc new_data: chỉ giữ ngày chưa có trong file
    rows_to_add = []
    for date_str, row_data in new_data.items():
        d = _parse_date(date_str)
        if d is None:
            if verbose:
                print(f"[!] Bỏ qua, không parse được ngày: {date_str}")
            continue
        if d.date() in existing_dates:
            if verbose:
                print(f"[-] Skip {date_str} (đã tồn tại trong file)")
            continue
        rows_to_add.append((d, row_data))

    if not rows_to_add:
        if verbose:
            print("[=] Không có ngày mới nào cần thêm.")
        wb.close()
        return 0

    # 5. Sắp xếp ngày mới theo thứ tự CŨ -> MỚI, rồi insert lần lượt vào ngay
    #    sau header (row header_row+1) -> kết quả cuối cùng ngày MỚI NHẤT
    #    sẽ nằm trên cùng, giống định dạng hiện tại của file.
    rows_to_add.sort(key=lambda x: x[0])  # cũ -> mới

    insert_at = header_row + 1
    for d, row_data in rows_to_add:
        ws.insert_rows(insert_at)
        date_cell = ws.cell(row=insert_at, column=date_col_idx, value=d)
        date_cell.font = body_font
        date_cell.number_format = date_excel_format
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        for fund_name, price in row_data.items():
            col_idx = headers[fund_name]
            cell = ws.cell(row=insert_at, column=col_idx, value=price)
            cell.font = body_font
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if verbose:
            print(f"[+] Đã thêm dòng {d.strftime(date_format)}")

    wb.save(file_path)
    wb.close()
    return len(rows_to_add)


if __name__ == "__main__":
    # Ví dụ test nhanh
    sample_new_data = {
        "25/06/2026": {
            "Quỹ Tăng trưởng": 10900.12,
            "Quỹ Cổ phần": 13150.00,
        }
    }
    added = update_fund_excel(
        file_path="gia_don_vi_quy_hanwha.xlsx",
        new_data=sample_new_data,
    )
    print(f"Đã thêm {added} dòng mới.")