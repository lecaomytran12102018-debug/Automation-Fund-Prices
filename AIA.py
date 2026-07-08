# -*- coding: utf-8 -*-
import re
import sys
import time
import unicodedata

from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from excel_fund_updater import update_fund_excel


URL = "https://www.aia.com.vn/vi/san-pham/lai-suat-va-gia-don-vi-quy.html"
EXCEL_FILE = "gia_don_vi_quy_TONGHOP.xlsx"  # file Excel gộp chung, đổi path nếu cần
SHEET_NAME = "AIA"  # sheet riêng cho AIA trong file gộp
DATE_COL_HEADER = "Ngày"

DATE_CELL_RE = re.compile(r"(\d{2}/\d{2}/\d{4})")
# Giá AIA có dấu chấm ngăn nghìn + hậu tố "VNĐ" dính liền, vd "16.887 VNĐ"
PRICE_CELL_RE = re.compile(r"([\d]{1,3}(?:\.\d{3})*)\s*VN[ĐD]")


def _nfc(text):
    return unicodedata.normalize("NFC", text or "").strip()


def _vn_price_to_float(price_str):
    """'16.887' -> 16887.0"""
    return float(price_str.replace(".", ""))


def get_target_fund_names(excel_file, sheet_name, date_col_header=DATE_COL_HEADER):
    wb = load_workbook(excel_file, read_only=True)
    ws = wb[sheet_name]   # ✅ chọn đúng sheet

    headers = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None and str(val).strip() != date_col_header:
            headers.append(_nfc(str(val)))

    wb.close()
    return headers


def fetch_rendered_html(url=URL, headless=False, wait_seconds=25):
    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1366,900")

    driver = webdriver.Chrome(options=options)
    try:
        driver.get(url)

        # ✅ scroll để trigger lazy load
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1)

        # ✅ wait cho data load đủ
        def wait_for_full_data(driver):
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_all_elements_located(
                        (By.XPATH, "//*[contains(text(),'VNĐ')]")
                    )
                )
                WebDriverWait(driver, 20).until(
                    EC.presence_of_all_elements_located(
                        (By.XPATH, "//*[contains(text(),'/')]")
                    )
                )
                return True
            except:
                return False

        wait_for_full_data(driver)
        time.sleep(2)

        return driver.page_source

    finally:
        driver.quit()    

    


def parse_fund_prices(html, target_fund_names):
    """
    Trả về: (date_str, {fund_col_name: price})

    Thử parse theo cấu trúc bảng thật (<table>/<tr>/<td>) trước. Nếu không
    tìm thấy <tr> nào (bảng render bằng div/grid), fallback sang cách quét
    text thô bằng cách định vị TỪNG TÊN QUỸ ĐÃ BIẾT TRƯỚC (lấy từ Excel) rồi
    tìm giá + ngày ngay sau đó -> không phụ thuộc đúng/sai cấu trúc HTML.
    """
    soup = BeautifulSoup(html, "html.parser")
    prices = {}
    found_date = None

    rows = soup.find_all("tr")
    for tr in rows:
        cells = tr.find_all(["td", "th"])
        if len(cells) < 2:
            continue

        cell_texts = [_nfc(c.get_text(separator=" ", strip=True)) for c in cells]
        first_cell = cell_texts[0]

        matched_fund = next((f for f in target_fund_names if f in first_cell), None)
        if not matched_fund:
            continue  # không phải dòng dữ liệu quỹ (có thể là header)

        # Bỏ qua hoàn toàn cột "Mức độ rủi ro" -> chỉ quét các ô còn lại để
        # tìm giá (có hậu tố VNĐ) và ngày (dd/mm/yyyy) theo đúng format.
        date_str = None
        price_val = None
        for text in cell_texts[1:]:
            if date_str is None:
                m = DATE_CELL_RE.search(text)
                if m:
                    date_str = m.group(1)
            if price_val is None:
                m = PRICE_CELL_RE.search(text)
                if m:
                    price_val = _vn_price_to_float(m.group(1))

        if date_str and price_val is not None:
            prices[matched_fund] = price_val
            found_date = date_str
        else:
            print(f"[!] Bỏ qua dòng '{matched_fund}': không tìm đủ ngày/giá trong ô: {cell_texts}")

    if prices:
        return found_date, prices

    # ---- Fallback: bảng không dùng <table> thật, quét text thô ----
    print("[!] Không tìm thấy dữ liệu qua <tr>, thử fallback quét text thô theo tên quỹ đã biết...")
    text = re.sub(r"<[^>]+>", " ", html)
    text = _nfc(re.sub(r"\s+", " ", text))

    for fund_name in target_fund_names:
        idx = text.find(fund_name)
        if idx == -1:
            continue
        # chỉ xét đoạn text ngay sau tên quỹ (trong khoảng 200 ký tự) để
        # tránh dính sang dòng của quỹ kế tiếp.
        chunk = text[idx: idx + 500]
        price_m = PRICE_CELL_RE.search(chunk)
        date_m = DATE_CELL_RE.search(chunk)
        if price_m and date_m:
            prices[fund_name] = _vn_price_to_float(price_m.group(1))
            found_date = date_m.group(1)
        else:
            print(f"[!] (fallback) Bỏ qua '{fund_name}': không tìm đủ ngày/giá trong đoạn: {chunk!r}")

    return found_date, prices


def main():
    target_funds = get_target_fund_names(EXCEL_FILE, SHEET_NAME)
    print(f"[*] Các quỹ cần tìm (lấy từ header Excel): {target_funds}")

    print(f"[*] Đang mở trình duyệt và fetch dữ liệu từ AIA ...")
    try:
        html = fetch_rendered_html(headless=False)
    except Exception as e:
        print(f"[x] Lỗi khi mở trang bằng Selenium: {e}")
        sys.exit(1)

    date_str, prices = parse_fund_prices(html, target_funds)

    if not date_str or not prices:
        print(
            "[x] Không parse được dữ liệu giá. Có thể cấu trúc trang đã đổi, "
            "hoặc trang chưa kịp render xong trước khi đọc."
        )
        sys.exit(1)

    missing = set(target_funds) - set(prices.keys())
    if missing:
        print(f"[!] Cảnh báo: thiếu giá các quỹ: {missing}")

    print(f"[*] Ngày định giá: {date_str}")
    for fund, price in prices.items():
        print(f"    {fund}: {price}")

    added = update_fund_excel(
        file_path=EXCEL_FILE,
        sheet_name=SHEET_NAME,
        new_data={date_str: prices},
    )

    if added:
        print(f"[+] Đã thêm {added} dòng mới vào {EXCEL_FILE}")
    else:
        print(f"[=] Không có dòng mới (ngày {date_str} đã tồn tại trong file).")


if __name__ == "__main__":
    main()